import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string


def automate_excel(file_name):
    file_excel = pd.read_excel(file_name)
    table_pivot = file_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    month_extension = file_name.split('_')[1]
    table_pivot.to_excel(f'sales_{month_extension}', startrow=4, sheet_name='Report')

    wb = load_workbook(f'sales_{month_extension}')
    tab = wb['Report']

    min_col = wb.active.min_column
    max_col = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    # graphic

    barChart = BarChart()

    data = Reference(tab, min_col=min_col + 1, max_col=max_col, min_row=min_row, max_row=max_row)
    category = Reference(tab, min_col=min_col, max_col=min_col, min_row=min_row + 1, max_row=max_row)

    barChart.add_data(data, titles_from_data=True)
    barChart.set_categories(category)

    tab.add_chart(barChart, 'B12')
    barChart.title = 'Ventas'
    barChart.style = 5

    abc = string.ascii_uppercase
    abc_excel = abc[0:max_col]

    for i in abc_excel:
        if i != 'A':
            tab[f'{i}{max_row + 1}'] = f'=SUM({i}{min_row + 1}:{i}{max_row})'
            tab[f'{i}{max_row + 1}'].style = 'Currency'

    tab[f'{abc_excel[0]}{max_col + 1}'] = 'Total'

    tab['A1'] = 'Reporte'
    month = month_extension.split('.')[0]
    tab['A2'] = month

    tab['A1'].font = Font('Arial', bold=True, size=20)
    tab['A2'].font = Font('Arial', bold=True, size=12)

    wb.save(f'sales_{month_extension}')
    return


automate_excel('supermarket_sales.xlsx')
